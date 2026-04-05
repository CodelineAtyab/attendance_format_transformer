"""
Generates opal_batch1_agile_oracles_attendance_march_2026.xlsx
from source_format_with_data.xlsm (March sheet).
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ── constants ────────────────────────────────────────────────────────────────
MONTH_NAME   = "March"
YEAR         = 2026
MONTH_NUM    = 3
TOTAL_DAYS   = 31          # days in March
FIRST_DAY    = date(YEAR, MONTH_NUM, 1)

# Determined empirically from source data:
# days where ALL employees have None  → weekend
# days where ALL employees have '-'   → public holiday (Eid Al-Fitr 2026)
WEEKEND_DAYS    = {6, 7, 13, 14, 20, 21, 28}   # all-None in source
PUBLIC_HOLIDAYS = {19, 22, 23}                   # all-dash in source

def build_day_types():
    """Returns a dict: day_num -> 'W' | 'PH' | 'work' """
    result = {}
    for d in range(1, TOTAL_DAYS + 1):
        if d in WEEKEND_DAYS:
            result[d] = "W"
        elif d in PUBLIC_HOLIDAYS:
            result[d] = "PH"
        else:
            result[d] = "work"
    return result

DAY_TYPES = build_day_types()
ACTUAL_WORKING_DAYS = sum(1 for t in DAY_TYPES.values() if t == "work")

# ── source code mapping ───────────────────────────────────────────────────────
def map_code(src_val, day_num):
    """Map a source attendance code to the destination format code."""
    if src_val is None:
        return DAY_TYPES[day_num] if DAY_TYPES[day_num] in ("W", "PH") else ""
    v = str(src_val).strip()
    if v == "-":
        return "PH"
    if v == "P":
        return "P"
    if v == "A":
        return "A"
    if v == "AP":
        return "OA"   # Absent with Permission → Official Absence
    if v == "SL":
        return "SL"
    if v == "L":
        return "AL"   # Leave (full day) → Annual Leave
    if v.startswith("L--"):
        return "P"    # Late but present
    return v          # pass through unknown values unchanged

# ── colours / styles ─────────────────────────────────────────────────────────
HEADER_BG   = "0066CC"   # blue header row
TITLE_BG    = "A7D9E7"   # light blue title
TITLE_FG    = "003366"   # dark navy title text
WHITE       = "FFFFFF"
WEEKEND_BG  = "D9D9D9"   # grey weekend cells
HOLIDAY_BG  = "FFE699"   # light yellow holiday cells
PRESENT_BG  = "C6EFCE"   # light green present
ABSENT_BG   = "FFC7CE"   # light red absent

def hdr_font(size=12, bold=True, color=WHITE):
    return Font(bold=bold, size=size, color=color)

def cell_font(size=11, bold=False, color="000000"):
    return Font(bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

# cell-colour based on destination code
CELL_FILLS = {
    "P":  "C6EFCE",  # green
    "A":  "FFC7CE",  # red
    "OA": "FFEB9C",  # yellow
    "SL": "BDD7EE",  # blue
    "AL": "E2EFDA",  # light green
    "PH": "FFE699",  # gold
    "W":  "D9D9D9",  # grey
    "":   "FFFFFF",  # white
}

# ── key legend ───────────────────────────────────────────────────────────────
KEY_LEGEND = [
    ("P",  "Present"),
    ("W",  "Weekend"),
    ("AL", "Annual Leave"),
    ("R",  "Rotation/days off"),
    ("PH", "Public Holiday"),
    ("A",  "Absent with no reason"),
    ("SL", "Sick Leave"),
    ("OA", "Official Absence"),
]
IMPORTANT_NOTE = (
    "Important Note: if a trainee is absent for a valid reason "
    "(e.g., Attending a job interview) and provides prior notice or proof, "
    "the trainee will be marked as OA in the attendance record."
)

# ── column layout ─────────────────────────────────────────────────────────────
# Col A=1 N | B=2 Name | C=3 ID | D=4 GSM | E=5 Company
# F=6 … AK=36 dates (31 days)
# AL=37 Attendance Days | AM=38 Actual days | AN=39 Stipend | AO=40 Total
# AP=41 Notes | AQ=42 Key Code | AR=43 description
DATE_START_COL  = 6      # column F
DATE_END_COL    = DATE_START_COL + TOTAL_DAYS - 1   # = 36 = AK
ATT_DAYS_COL    = DATE_END_COL + 1   # AL
ACTUAL_DAYS_COL = DATE_END_COL + 2   # AM
STIPEND_COL     = DATE_END_COL + 3   # AN
TOTAL_COL       = DATE_END_COL + 4   # AO
NOTES_COL       = DATE_END_COL + 5   # AP
KEY_CODE_COL    = DATE_END_COL + 6   # AQ
KEY_DESC_COL    = DATE_END_COL + 7   # AR
TOTAL_COLS      = KEY_DESC_COL


def col_letter(n):
    return get_column_letter(n)


# ── read source data ──────────────────────────────────────────────────────────
def read_source_march():
    wb = openpyxl.load_workbook(
        "source_format_with_data.xlsm", keep_vba=True, data_only=True
    )
    ws = wb["March"]
    employees = []
    for r in range(7, ws.max_row + 1):
        emp_id = ws.cell(r, 1).value
        name   = ws.cell(r, 2).value
        if emp_id is None or name is None:
            continue
        # columns C(3) … AG(33) → days 1..31
        days_raw = [ws.cell(r, 2 + d).value for d in range(1, TOTAL_DAYS + 1)]
        employees.append({"id": int(emp_id), "name": name, "days_raw": days_raw})
    return employees


# ── write output workbook ─────────────────────────────────────────────────────
def write_output(employees):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTH_NAME} {YEAR} Attendance"

    # ── ROW 1 – title ─────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 55
    title_cell = ws.cell(1, 1)
    title_cell.value     = "Tamayuz Program Trainees Attendance Report"
    title_cell.font      = Font(bold=True, size=22, color=TITLE_FG)
    title_cell.fill      = fill(TITLE_BG)
    title_cell.alignment = center_align()
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1,   end_column=NOTES_COL
    )
    # side panel title
    kh = ws.cell(1, KEY_CODE_COL)
    kh.value     = "Attendance Key Codes"
    kh.font      = Font(bold=True, size=12, color=TITLE_FG)
    kh.fill      = fill(TITLE_BG)
    kh.alignment = center_align()
    ws.merge_cells(
        start_row=1, start_column=KEY_CODE_COL,
        end_row=1,   end_column=KEY_DESC_COL
    )

    # ── ROW 2 – column headers ────────────────────────────────────────────
    ws.row_dimensions[2].height = 40
    headers = [
        (1, "N"), (2, "Trainee Name"), (3, "ID"), (4, "GSM"), (5, "Company"),
    ]
    for col, val in headers:
        c = ws.cell(2, col)
        c.value = val; c.font = hdr_font(); c.fill = fill(HEADER_BG)
        c.alignment = center_align(wrap=True); c.border = thin_border()

    for d in range(1, TOTAL_DAYS + 1):
        col = DATE_START_COL + d - 1
        dt  = date(YEAR, MONTH_NUM, d)
        c   = ws.cell(2, col)
        c.value = dt
        c.number_format = "D\nMMM"   # e.g. "1\nMar"
        c.font      = hdr_font(size=10)
        c.fill      = fill(WEEKEND_BG if DAY_TYPES[d] == "W" else
                           HOLIDAY_BG if DAY_TYPES[d] == "PH" else HEADER_BG)
        c.alignment = center_align(wrap=True)
        c.border    = thin_border()

    for col, lbl in [
        (ATT_DAYS_COL,    "Attendance\nDays"),
        (ACTUAL_DAYS_COL, "Actual Days\nin Month"),
        (STIPEND_COL,     "Stipend Amount\nas per Contract"),
        (TOTAL_COL,       "Total Amount"),
        (NOTES_COL,       "Notes"),
    ]:
        c = ws.cell(2, col)
        c.value = lbl; c.font = hdr_font(size=10); c.fill = fill(HEADER_BG)
        c.alignment = center_align(wrap=True); c.border = thin_border()

    # key codes header (merged rows 2-3)
    kh2 = ws.cell(2, KEY_CODE_COL)
    kh2.value     = "Key\nCode"
    kh2.font      = hdr_font(size=10)
    kh2.fill      = fill(HEADER_BG)
    kh2.alignment = center_align(wrap=True)
    kh2.border    = thin_border()

    kh3 = ws.cell(2, KEY_DESC_COL)
    kh3.value     = "Description"
    kh3.font      = hdr_font(size=10)
    kh3.fill      = fill(HEADER_BG)
    kh3.alignment = center_align(wrap=True)
    kh3.border    = thin_border()

    # ── DATA ROWS ─────────────────────────────────────────────────────────
    for emp_idx, emp in enumerate(employees):
        row = 3 + emp_idx
        ws.row_dimensions[row].height = 15

        # N, Name, ID, GSM, Company
        for col, val in [
            (1, emp["id"]),
            (2, emp["name"]),
            (3, ""),   # ID not in source
            (4, ""),   # GSM not in source
            (5, ""),   # Company not in source
        ]:
            c = ws.cell(row, col)
            c.value = val; c.font = cell_font()
            c.alignment = center_align()
            c.border = thin_border()

        # daily attendance cells
        att_count = 0
        for d in range(1, TOTAL_DAYS + 1):
            col      = DATE_START_COL + d - 1
            src_val  = emp["days_raw"][d - 1]
            dest_val = map_code(src_val, d)
            if dest_val == "P":
                att_count += 1
            c = ws.cell(row, col)
            c.value = dest_val
            c.font  = cell_font(size=10)
            c.fill  = fill(CELL_FILLS.get(dest_val, "FFFFFF"))
            c.alignment = center_align()
            c.border = thin_border()

        # summary columns
        att_letter     = col_letter(ATT_DAYS_COL)
        actual_letter  = col_letter(ACTUAL_DAYS_COL)
        stipend_letter = col_letter(STIPEND_COL)

        ws.cell(row, ATT_DAYS_COL).value    = att_count
        ws.cell(row, ACTUAL_DAYS_COL).value = ACTUAL_WORKING_DAYS
        ws.cell(row, STIPEND_COL).value     = ""      # to be filled manually
        total_formula = (
            f"=IF({stipend_letter}{row}=\"\",\"\","
            f"{stipend_letter}{row}/{actual_letter}{row}*{att_letter}{row})"
        )
        ws.cell(row, TOTAL_COL).value = total_formula

        for col in [ATT_DAYS_COL, ACTUAL_DAYS_COL, STIPEND_COL, TOTAL_COL, NOTES_COL]:
            c = ws.cell(row, col)
            c.font      = cell_font()
            c.alignment = center_align()
            c.border    = thin_border()

        # key legend entries (right panel)
        if emp_idx < len(KEY_LEGEND):
            code, desc = KEY_LEGEND[emp_idx]
            kc = ws.cell(row, KEY_CODE_COL)
            kc.value = code; kc.font = cell_font(bold=True)
            kc.alignment = center_align(); kc.border = thin_border()
            kc.fill = fill(CELL_FILLS.get(code, "FFFFFF"))

            kd = ws.cell(row, KEY_DESC_COL)
            kd.value = desc; kd.font = cell_font()
            kd.alignment = Alignment(horizontal="left", vertical="center")
            kd.border = thin_border()

    # important note below legend
    note_row = 3 + len(KEY_LEGEND)
    note_end  = 3 + len(employees) - 1
    if note_row <= note_end:
        nc = ws.cell(note_row, KEY_CODE_COL)
        nc.value = IMPORTANT_NOTE
        nc.font  = Font(size=9, italic=True)
        nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(
            start_row=note_row, start_column=KEY_CODE_COL,
            end_row=note_end,   end_column=KEY_DESC_COL
        )

    # ── column widths ────────────────────────────────────────────────────
    ws.column_dimensions[col_letter(1)].width  = 5
    ws.column_dimensions[col_letter(2)].width  = 32
    ws.column_dimensions[col_letter(3)].width  = 12
    ws.column_dimensions[col_letter(4)].width  = 14
    ws.column_dimensions[col_letter(5)].width  = 16
    for d in range(1, TOTAL_DAYS + 1):
        ws.column_dimensions[col_letter(DATE_START_COL + d - 1)].width = 5.5
    ws.column_dimensions[col_letter(ATT_DAYS_COL)].width    = 12
    ws.column_dimensions[col_letter(ACTUAL_DAYS_COL)].width = 12
    ws.column_dimensions[col_letter(STIPEND_COL)].width     = 18
    ws.column_dimensions[col_letter(TOTAL_COL)].width       = 14
    ws.column_dimensions[col_letter(NOTES_COL)].width       = 18
    ws.column_dimensions[col_letter(KEY_CODE_COL)].width    = 8
    ws.column_dimensions[col_letter(KEY_DESC_COL)].width    = 36

    # ── freeze panes ─────────────────────────────────────────────────────
    ws.freeze_panes = "F3"

    # ── page setup ───────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1

    fname = f"opal_batch1_agile_oracles_attendance_march_{YEAR}.xlsx"
    wb.save(fname)
    return fname


# ── main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    employees = read_source_march()
    print(f"Loaded {len(employees)} employees from March sheet.")
    print(f"March {YEAR}: {TOTAL_DAYS} days | "
          f"Working days: {ACTUAL_WORKING_DAYS} | "
          f"Public holidays: {sorted(PUBLIC_HOLIDAYS)}")

    fname = write_output(employees)
    print(f"Output saved: {fname}")

    # quick summary
    print("\nAttendance summary:")
    print(f"{'#':<4} {'Name':<35} {'Present':>8} {'Actual':>8}")
    print("-" * 58)
    for emp in employees:
        att = sum(
            1 for d, v in enumerate(emp["days_raw"], 1)
            if map_code(v, d) == "P"
        )
        print(f"{emp['id']:<4} {emp['name']:<35} {att:>8} {ACTUAL_WORKING_DAYS:>8}")
